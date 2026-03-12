"""Build TEST_CASES_Laundry_POS.xlsx with color coding: green = Normal, red = Error Handling."""
import csv
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CSV_PATH = os.path.join(SCRIPT_DIR, "TEST_CASES_Laundry_POS.csv")
XLSX_PATH = os.path.join(SCRIPT_DIR, "TEST_CASES_Laundry_POS.xlsx")

# Colors: green for normal, red for error handling (light shades for readability)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # Light green
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Light red
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FONT = Font(bold=True)

wb = Workbook()
ws = wb.active
ws.title = "Test Cases"

with open(CSV_PATH, "r", encoding="utf-8") as f:
    reader = csv.reader(f)
    rows = list(reader)

# Type column index (second column = index 1)
TYPE_COL = 1

for row_idx, row in enumerate(rows, start=1):
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        else:
            # Color by Type: Normal = green, Error Handling = red
            if len(row) > TYPE_COL and row[TYPE_COL].strip() == "Error Handling":
                cell.fill = RED_FILL
            else:
                cell.fill = GREEN_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# Adjust column widths for readability
for col_idx in range(1, len(rows[0]) + 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = 18 if col_idx <= 2 else 40

wb.save(XLSX_PATH)
print(f"Created: {XLSX_PATH}")
print("Color coding: Green = Normal test cases, Red = Error Handling test cases.")
